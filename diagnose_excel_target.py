from __future__ import annotations

import sys
import time
import unicodedata
from pathlib import Path
from zipfile import BadZipFile

from app.imei_normalizer import IMEI_SEPARATOR_CHARS, is_target_row, normalize_imei

try:
    import openpyxl
    from openpyxl.utils.exceptions import InvalidFileException
except Exception:  # pragma: no cover
    openpyxl = None

    class InvalidFileException(Exception):
        pass

from app.config import load_config
from app.excel_reader import ExcelReader
from app.excel_session import (
    ReadOnlyWorkbookError,
    SaveCloseWorkbookError,
    _resolve_web_identity_hash,
    detect_target_workbook,
    save_and_close_target_workbook,
)
from app.logger import AppLogger
from app.main import reopen_excel


class UnlockTimeoutError(RuntimeError):
    pass


class ReopenFailedError(RuntimeError):
    pass


IMEI_FORMAT_COUNTERS = (
    "total_candidate_row_count",
    "valid_original_count",
    "valid_after_whitespace_normalization_count",
    "blank_imei_count",
    "invalid_length_count",
    "non_ascii_digit_count",
    "non_digit_character_count",
    "non_integer_numeric_count",
    "invisible_character_count",
    "normalization_failure_count",
)
def _base_dir() -> Path:
    return Path(__file__).resolve().parent


def _is_valid_imei(value: str) -> bool:
    return len(value) == 15 and value.isascii() and value.isdigit()


def _new_imei_format_counts() -> dict[str, int]:
    return {key: 0 for key in IMEI_FORMAT_COUNTERS}


def _normalize_imei_for_diagnosis(value) -> tuple[str, str]:
    try:
        original = unicodedata.normalize("NFC", "" if value is None else str(value))
    except Exception:
        return "", "normalization_failure_count"

    if not original or all(char in IMEI_SEPARATOR_CHARS for char in original):
        return "", "blank_imei_count"
    try:
        result = normalize_imei(value)
    except ValueError as exc:
        if str(exc) == "IMEIが整数ではありません":
            return "", "non_integer_numeric_count"
        if any(
            unicodedata.category(char) == "Cf" and char not in IMEI_SEPARATOR_CHARS
            for char in original
        ):
            return "", "invisible_character_count"
        if any(
            char not in IMEI_SEPARATOR_CHARS
            and char.isdigit()
            and not ("0" <= char <= "9")
            for char in original
        ):
            return "", "non_ascii_digit_count"
        if str(exc) == "IMEIは15桁ではありません":
            return "", "invalid_length_count"
        return "", "non_digit_character_count"
    if isinstance(value, (int, float)) or _is_valid_imei(original):
        return result, "valid_original_count"
    return result, "valid_after_whitespace_normalization_count"


def _is_candidate_row(values: tuple[object, ...]) -> bool:
    return any(value is not None and str(value) != "" for value in values)


def _collect_imei_format_counts(excel_path: Path) -> dict[str, int]:
    if openpyxl is None:
        raise ImportError("openpyxl unavailable")

    counts = _new_imei_format_counts()
    workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True, keep_vba=True)
    try:
        if "HENNGE登録作業必要情報" not in workbook.sheetnames:
            raise KeyError("sheet not found")
        sheet = workbook["HENNGE登録作業必要情報"]
        for row in sheet.iter_rows(min_row=4, max_col=5, values_only=True):
            values = tuple(row[2:5])
            if not is_target_row(values):
                continue
            counts["total_candidate_row_count"] += 1
            _normalized, category = _normalize_imei_for_diagnosis(values[2] if len(values) > 2 else None)
            counts[category] += 1
        return counts
    finally:
        workbook.close()


def _emit_imei_format_counts(logger: AppLogger, counts: dict[str, int]) -> None:
    for key in IMEI_FORMAT_COUNTERS:
        _emit_kv(logger, key, counts[key])


def _emit_kv(logger: AppLogger, key: str, value) -> None:
    logger.info(f"{key}={value}")


def _is_sheet_not_found_error(exc: Exception) -> bool:
    if isinstance(exc, KeyError):
        return True
    message = str(exc).lower()
    return "sheet" in message and ("not found" in message or "missing" in message)


def _log_failure(logger: AppLogger, stage: str, exc: BaseException) -> None:
    logger.error(f"failed_stage={stage}")
    logger.error(f"exception_type={type(exc).__name__}")


def _classify_value_error(exc: ValueError) -> tuple[str, int]:
    msg = str(exc)
    if msg.startswith("Excel ") and "行目の必須項目が不足: " in msg:
        return "missing_required_field", 10
    if msg in {
        "IMEIが整数ではありません",
        "IMEIに数字以外が含まれています",
        "IMEIは15桁ではありません",
    }:
        return "invalid_imei", 5
    return "unknown_value_error", 1


def _wait_unlock(reader: ExcelReader, timeout_sec: float = 15.0, interval_sec: float = 0.5) -> None:
    deadline = time.monotonic() + timeout_sec
    while time.monotonic() < deadline:
        if not reader.is_file_open():
            return
        time.sleep(interval_sec)
    raise UnlockTimeoutError("unlock timeout")


class _NoopLogger:
    def info(self, _message: str) -> None:
        return None


def _run_diagnostic(argv: list[str] | None = None) -> int:
    args = argv if argv is not None else sys.argv[1:]
    format_mode = args == ["--diagnose-imei-format"]
    logger = AppLogger(_base_dir())
    current_stage = "init"
    reopen_path: Path | None = None
    reopen_required = False
    reopen_attempted = False

    def _reopen_once() -> None:
        nonlocal reopen_attempted
        if reopen_attempted or reopen_path is None:
            return
        reopen_attempted = True
        signal = {"failed": False}

        def _capture_reopen_result(message: str) -> None:
            if "起動に失敗" in message:
                signal["failed"] = True

        reopen_excel(reopen_path, _capture_reopen_result)
        if not format_mode:
            _emit_kv(logger, "reopen_called", True)
        if signal["failed"]:
            raise RuntimeError("reopen failed")

    try:
        current_stage = "config_loaded"
        config = load_config()

        current_stage = "excel_path_resolved"
        excel_path = (config.get("excel", {}) or {}).get("path", "")
        if not excel_path:
            return 8
        reopen_path = Path(excel_path)
        if not reopen_path.exists():
            return 8

        current_stage = "reader_created"
        reader = ExcelReader(excel_path)

        current_stage = "file_open_check_completed"
        file_open = reader.is_file_open()
        if not format_mode:
            _emit_kv(logger, "target_was_open", file_open)
        if file_open:
            excel_config = config.get("excel", {}) or {}
            web_identity = _resolve_web_identity_hash(excel_config, "test")

            current_stage = "workbook_detection"
            detection = detect_target_workbook(
                reopen_path.resolve(),
                configured_web_identity_hash=web_identity.hash_value,
                web_identity_source=web_identity.source,
                web_identity_valid=web_identity.valid,
                web_identity_mode="test",
                timeout_sec=15.0,
                interval_sec=0.5,
            )
            if not format_mode:
                _emit_kv(logger, "matched_workbook_count", detection.matched_workbook_count)
                _emit_kv(logger, "target_match_method", detection.target_match_method)
            if detection.workbook is None or detection.application is None or detection.matched_workbook_count != 1:
                if not format_mode:
                    _emit_kv(logger, "save_called", False)
                    _emit_kv(logger, "close_called", False)
                    _emit_kv(logger, "unlock_completed", False)
                    _emit_kv(logger, "read_targets_called", False)
                return 8 if format_mode else 9

            current_stage = "save_close"
            if not format_mode:
                _emit_kv(logger, "save_called", True)
            saved_and_closed = save_and_close_target_workbook(
                reopen_path.resolve(),
                _NoopLogger(),
                configured_web_identity_hash=web_identity.hash_value,
                web_identity_source=web_identity.source,
                web_identity_valid=web_identity.valid,
                web_identity_mode="test",
            )
            if not saved_and_closed:
                if not format_mode:
                    _emit_kv(logger, "close_called", False)
                    _emit_kv(logger, "unlock_completed", False)
                    _emit_kv(logger, "read_targets_called", False)
                return 1 if format_mode else 12
            if not format_mode:
                _emit_kv(logger, "close_called", True)
            reopen_required = True

            current_stage = "unlock_wait"
            _wait_unlock(reader, timeout_sec=15.0, interval_sec=0.5)
            if not format_mode:
                _emit_kv(logger, "unlock_completed", True)
        else:
            if not format_mode:
                _emit_kv(logger, "matched_workbook_count", 0)
                _emit_kv(logger, "target_match_method", "none")
                _emit_kv(logger, "save_called", False)
                _emit_kv(logger, "close_called", False)
                _emit_kv(logger, "unlock_completed", True)

        if format_mode:
            current_stage = "imei_format_scan_started"
            counts = _collect_imei_format_counts(reopen_path)
            _emit_imei_format_counts(logger, counts)
            invalid_count = counts["total_candidate_row_count"] - (
                counts["valid_original_count"]
                + counts["valid_after_whitespace_normalization_count"]
            )
            if counts["total_candidate_row_count"] == 0:
                return 2
            return 5 if invalid_count > 0 else 0

        current_stage = "read_targets_started"
        _emit_kv(logger, "read_targets_called", True)
        targets = reader.read_targets()
        current_stage = "read_targets_completed"

        current_stage = "target_validation_started"
        target_count = len(targets)
        _emit_kv(logger, "target_count", target_count)
        if target_count == 0:
            return 2

        validated_rows: list[dict] = []
        for row in targets:
            if not isinstance(row, dict):
                return 11
            alias = str(row.get("alias", "")).strip()
            imei = str(row.get("imei", "")).strip()
            if not alias or not imei:
                return 10
            if not _is_valid_imei(imei):
                return 5
            validated_rows.append({"alias": alias, "imei": imei})

        aliases = [row["alias"] for row in validated_rows]
        imeis = [row["imei"] for row in validated_rows]
        if len(set(aliases)) != len(aliases) or len(set(imeis)) != len(imeis):
            return 11

        return 0
    except KeyboardInterrupt:
        if not format_mode:
            _log_failure(logger, current_stage, KeyboardInterrupt())
        return 130
    except (ReadOnlyWorkbookError, SaveCloseWorkbookError) as exc:
        if not format_mode:
            _log_failure(logger, current_stage, exc)
        return 1 if format_mode else 12
    except UnlockTimeoutError as exc:
        if not format_mode:
            _emit_kv(logger, "unlock_completed", False)
            _emit_kv(logger, "read_targets_called", False)
            _log_failure(logger, current_stage, exc)
        return 1 if format_mode else 13
    except ValueError as exc:
        category, code = _classify_value_error(exc)
        if not format_mode:
            _log_failure(logger, current_stage, exc)
            _emit_kv(logger, "error_category", category)
        return 1 if format_mode else code
    except Exception as exc:
        if not format_mode:
            _log_failure(logger, current_stage, exc)
        if isinstance(exc, (FileNotFoundError, KeyError)) or _is_sheet_not_found_error(exc):
            return 8
        if isinstance(exc, (PermissionError, InvalidFileException, BadZipFile, TypeError)):
            return 1
        return 1
    finally:
        if reopen_required and not reopen_attempted:
            try:
                current_stage = "reopen"
                _reopen_once()
            except Exception as exc:
                if not format_mode:
                    _log_failure(logger, current_stage, exc)
                raise ReopenFailedError("reopen failed") from None


def main(argv: list[str] | None = None) -> int:
    try:
        return _run_diagnostic(argv)
    except ReopenFailedError:
        args = argv if argv is not None else sys.argv[1:]
        return 1 if args == ["--diagnose-imei-format"] else 14


if __name__ == "__main__":
    raise SystemExit(main())
