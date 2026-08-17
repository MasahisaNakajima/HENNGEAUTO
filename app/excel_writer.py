from __future__ import annotations

from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:
    openpyxl = None


class ExcelWriter:
    def __init__(self, file_path: str | Path, *, sheet_name: str = "HENNGE登録作業必要情報", column: int = 6, value: Any = "完了") -> None:
        self.file_path = Path(file_path)
        self.sheet_name = sheet_name
        self.column = int(column)
        self.value = value

    def plan_update(self, *, row_number: int, alias: str, serial: str, imei: str) -> dict[str, Any]:
        if row_number < 1 or not alias or not serial or not imei:
            raise ValueError("Excel更新対象を確認できません")
        return {"row_number": row_number, "column": self.column, "source_match_required": True}

    def write_update(self, *, row_number: int, alias: str, serial: str, imei: str) -> dict[str, Any]:
        if openpyxl is None:
            raise ImportError("openpyxl が未インストールです")
        self.plan_update(row_number=row_number, alias=alias, serial=serial, imei=imei)
        workbook = openpyxl.load_workbook(self.file_path, keep_vba=self.file_path.suffix.casefold() == ".xlsm")
        try:
            if self.sheet_name not in workbook.sheetnames:
                raise KeyError("対象シートが見つかりません")
            sheet = workbook[self.sheet_name]
            if str(sheet.cell(row_number, 3).value or "").strip() != alias:
                raise RuntimeError("Excel alias照合に失敗しました")
            if str(sheet.cell(row_number, 4).value or "").strip() != serial:
                raise RuntimeError("Excel serial照合に失敗しました")
            if str(sheet.cell(row_number, 5).value or "").strip() != imei:
                raise RuntimeError("Excel IMEI照合に失敗しました")
            sheet.cell(row_number, self.column).value = self.value
            workbook.save(self.file_path)
            return {"row_number": row_number, "column": self.column, "written": True}
        finally:
            workbook.close()
