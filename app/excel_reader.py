import warnings
from pathlib import Path

from app.imei_normalizer import normalize_imei
from app.imei_normalizer import is_target_row

try:
    import openpyxl
except ImportError:
    openpyxl = None


class ExcelReader:
    def __init__(self, file_path: str, trace_callback=None):
        self.file_path = Path(file_path)
        self.trace_callback = trace_callback
        self.trace_counters = {
            "normalizer_invoked": 0,
            "normalizer_success_count": 0,
            "normalizer_failure_count": 0,
            "candidate_row_count": 0,
            "completed_row_count": 0,
        }

    def _trace(self, stage: str) -> None:
        if self.trace_callback is not None:
            self.trace_callback(stage)

    def read_targets(
        self,
        sheet_name: str = "HENNGE登録作業必要情報",
        *,
        include_row_number: bool = False,
    ) -> list[dict]:
        if openpyxl is None:
            raise ImportError("openpyxl が未インストールです。requirements.txt を確認してください。")

        warnings.filterwarnings(
            "ignore",
            message=r"Cell .* is marked as a date but the serial value .* is outside the limits for dates.*",
            category=UserWarning,
        )
        self._trace("workbook_load_started")
        workbook = openpyxl.load_workbook(self.file_path, data_only=True)
        try:
            self._trace("workbook_load_completed")
            self._trace("sheet_lookup_started")
            if sheet_name not in workbook.sheetnames:
                raise KeyError(f"対象シートが見つかりません: {sheet_name}")

            sheet = workbook[sheet_name]
            self._trace("sheet_lookup_completed")
            rows = []
            self._trace("row_iteration_started")
            for row_number, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
                alias = self._normalize_value(row[2]) if len(row) > 2 else ""
                serial = self._normalize_value(row[3]) if len(row) > 3 else ""

                if not is_target_row(tuple(row[2:5])):
                    continue
                self.trace_counters["candidate_row_count"] += 1

                self._trace("imei_normalization_started")
                self.trace_counters["normalizer_invoked"] += 1
                try:
                    imei = normalize_imei(row[4]) if len(row) > 4 else ""
                except ValueError:
                    self.trace_counters["normalizer_failure_count"] += 1
                    raise
                self.trace_counters["normalizer_success_count"] += 1
                self._trace("imei_normalization_completed")

                self._trace("required_field_validation_started")
                missing = [
                    name for name, value in {
                        "alias": alias,
                        "serial": serial,
                        "imei": imei,
                    }.items() if not value
                ]
                if missing:
                    raise ValueError(f"Excel {row_number}行目の必須項目が不足: {', '.join(missing)}")
                self._trace("required_field_validation_completed")

                self._trace("row_append_started")
                target = {
                    "alias": alias,
                    "serial": serial,
                    "imei": imei,
                }
                if include_row_number:
                    target["row_number"] = row_number
                rows.append(target)
                self.trace_counters["completed_row_count"] += 1
                self._trace("row_append_completed")

            self._trace("read_targets_returned")
            return rows
        finally:
            workbook.close()

    def is_file_open(self) -> bool:
        try:
            with self.file_path.open("rb+"):
                return False
        except PermissionError:
            return True

    @staticmethod
    def _normalize_value(value) -> str:
        return "" if value is None else str(value).strip()

    @staticmethod
    def _is_header_row(alias: str, serial: str, imei: str) -> bool:
        alias_key = alias.replace(" ", "").upper()
        serial_key = serial.replace(" ", "").upper()
        imei_key = imei.replace(" ", "").upper()
        return (
            alias_key in {"ALIAS", "エイリアス"}
            or serial_key in {"SERIAL", "シリアル", "シリアル番号"}
            or imei_key == "IMEI"
        )

    @staticmethod
    def _is_placeholder_row(alias: str, serial: str, imei: str) -> bool:
        markers = {"#N/A", "N/A", "NA", "-", ""}
        alias_key = alias.strip().upper()
        serial_key = serial.strip().upper()
        imei_key = imei.strip().upper()
        return alias_key in markers and serial_key in markers and imei_key in markers
