from __future__ import annotations

from pathlib import Path
from types import SimpleNamespace
from zipfile import BadZipFile

import pytest
from openpyxl import Workbook

import diagnose_excel_target as mod
from app.excel_session import WEB_IDENTITY_ENV_TEST
from app.excel_reader import ExcelReader


class DummyLogger:
    def __init__(self):
        self.info_messages = []
        self.error_messages = []

    def info(self, message):
        self.info_messages.append(message)

    def error(self, message):
        self.error_messages.append(message)


class FakeReader:
    def __init__(self, file_path, rows, calls, *, is_open=False, read_exc=None):
        self.file_path = file_path
        self._rows = rows
        self._calls = calls
        self._is_open = is_open
        self._read_exc = read_exc

    def is_file_open(self):
        self._calls["is_file_open"] += 1
        self._calls["events"].append("is_file_open")
        return self._is_open

    def read_targets(self):
        self._calls["read_targets"] += 1
        self._calls["events"].append("read_targets")
        if self._read_exc is not None:
            raise self._read_exc
        return list(self._rows)


def _install(
    monkeypatch,
    *,
    rows,
    config_path="X:/【テスト用】targets.xlsm",
    is_open=False,
    read_exc=None,
    detection_match=True,
    save_exc=None,
    close_exc=None,
    save_result=True,
    unlock_exc=None,
    reopen_fails=False,
):
    logger = DummyLogger()
    calls = {
        "load_config": 0,
        "reader_ctor": 0,
        "reader_path": "",
        "is_file_open": 0,
        "read_targets": 0,
        "detect": 0,
        "save": 0,
        "close": 0,
        "unlock": 0,
        "reopen": 0,
        "events": [],
        "detection_kwargs": {},
        "save_kwargs": {},
    }

    monkeypatch.setattr(mod, "AppLogger", lambda _base_dir: logger)

    def fake_load_config():
        calls["load_config"] += 1
        return {"excel": {"path": config_path}}

    def fake_reader_ctor(file_path):
        calls["reader_ctor"] += 1
        calls["reader_path"] = file_path
        return FakeReader(file_path, rows, calls, is_open=is_open, read_exc=read_exc)

    monkeypatch.setattr(mod, "load_config", fake_load_config)
    monkeypatch.setattr(mod, "ExcelReader", fake_reader_ctor)

    def fake_detect(*_args, **kwargs):
        calls["detect"] += 1
        calls["events"].append("detect")
        calls["detection_kwargs"] = kwargs
        return SimpleNamespace(
            workbook=object() if detection_match else None,
            application=object() if detection_match else None,
            matched_workbook_count=1 if detection_match else 0,
            target_match_method="web_identity" if detection_match else "none",
        )

    def fake_save_close(*_args, **kwargs):
        calls["save"] += 1
        calls["events"].append("save")
        calls["save_kwargs"] = kwargs
        if save_exc is not None:
            raise save_exc
        if save_result:
            calls["close"] += 1
            calls["events"].append("close")
            if close_exc is not None:
                raise close_exc
        return save_result

    def fake_wait_unlock(*_args, **_kwargs):
        calls["unlock"] += 1
        calls["events"].append("unlock")
        if unlock_exc is not None:
            raise unlock_exc

    def fake_reopen(_path, emit):
        calls["reopen"] += 1
        calls["events"].append("reopen")
        if reopen_fails:
            emit("Excelファイルの起動に失敗しました: RuntimeError")
        else:
            emit("Excelファイルを起動しました")

    monkeypatch.setattr(mod.Path, "exists", lambda _self: True)
    monkeypatch.setattr(mod, "detect_target_workbook", fake_detect)
    monkeypatch.setattr(mod, "save_and_close_target_workbook", fake_save_close)
    monkeypatch.setattr(mod, "_wait_unlock", fake_wait_unlock)
    monkeypatch.setattr(mod, "reopen_excel", fake_reopen)
    return logger, calls


def test_reuses_excel_reader_and_main_config_path(monkeypatch):
    rows = [{"alias": "alias01", "serial": "s1", "imei": "123456789012345"}]
    logger, calls = _install(monkeypatch, rows=rows, config_path="C:/excel/live.xlsm")

    rc = mod.main([])

    assert rc == 0
    assert calls["load_config"] == 1
    assert calls["reader_ctor"] == 1
    assert calls["is_file_open"] == 1
    assert calls["read_targets"] == 1
    assert calls["reader_path"] == "C:/excel/live.xlsm"
    assert "target_count=1" in "\n".join(logger.info_messages)
    assert calls["save"] == 0
    assert calls["close"] == 0


def test_success_when_exactly_one_valid_target(monkeypatch):
    rows = [{"alias": "abxyz", "serial": "s1", "imei": "123456789012345"}]
    logger, _calls = _install(monkeypatch, rows=rows)

    rc = mod.main([])

    joined = "\n".join(logger.info_messages + logger.error_messages)
    assert rc == 0
    assert "target_count=1" in joined


def test_is_file_open_true_saves_closes_unlocks_then_reads_targets(monkeypatch):
    rows = [{"alias": "abxyz", "serial": "s1", "imei": "123456789012345"}]
    logger, calls = _install(monkeypatch, rows=rows, is_open=True)

    rc = mod.main([])

    assert rc == 0
    assert calls["is_file_open"] == 1
    assert calls["detect"] == 1
    assert calls["save"] == 1
    assert calls["close"] == 1
    assert calls["unlock"] == 1
    assert calls["read_targets"] == 1
    assert calls["reopen"] == 1
    assert calls["events"] == ["is_file_open", "detect", "save", "close", "unlock", "read_targets", "reopen"]
    joined = "\n".join(logger.info_messages)
    assert "target_was_open=True" in joined
    assert "matched_workbook_count=1" in joined
    assert "target_match_method=web_identity" in joined


def test_is_file_open_false_calls_read_targets(monkeypatch):
    rows = [{"alias": "abxyz", "serial": "s1", "imei": "123456789012345"}]
    _logger, calls = _install(monkeypatch, rows=rows, is_open=False)

    rc = mod.main([])

    assert rc == 0
    assert calls["is_file_open"] == 1
    assert calls["read_targets"] == 1


@pytest.mark.parametrize(
    "rows, expected",
    [([], 2), ([{"alias": "a1", "serial": "s1", "imei": "123456789012345"}, {"alias": "a2", "serial": "s2", "imei": "223456789012345"}], 0)],
)
def test_distinguishes_zero_and_multiple_targets(monkeypatch, rows, expected):
    _logger, _calls = _install(monkeypatch, rows=rows)

    rc = mod.main([])

    assert rc == expected


def test_keeps_alias_imei_as_same_row_pair(monkeypatch):
    rows = [
        {"alias": "pairA", "serial": "s1", "imei": "123456789012345"},
        {"alias": "pairB", "serial": "s2", "imei": "223456789012345"},
    ]
    observed = {"rows": []}

    class PairReader(FakeReader):
        def is_file_open(self):
            self._calls["is_file_open"] += 1
            return False

        def read_targets(self):
            observed["rows"] = list(self._rows)
            return super().read_targets()

    logger = DummyLogger()
    calls = {"read_targets": 0, "is_file_open": 0, "events": []}
    monkeypatch.setattr(mod, "AppLogger", lambda _base_dir: logger)
    monkeypatch.setattr(mod, "load_config", lambda: {"excel": {"path": "X:/pair.xlsm"}})
    monkeypatch.setattr(mod, "ExcelReader", lambda file_path: PairReader(file_path, rows, calls))
    monkeypatch.setattr(mod.Path, "exists", lambda _self: True)

    rc = mod.main([])

    assert rc == 0
    assert observed["rows"][0]["alias"] == "pairA"
    assert observed["rows"][0]["imei"] == "123456789012345"
    assert observed["rows"][1]["alias"] == "pairB"
    assert observed["rows"][1]["imei"] == "223456789012345"


@pytest.mark.parametrize("alias", ["", "   "])
def test_rejects_blank_alias(monkeypatch, alias):
    rows = [{"alias": alias, "serial": "s1", "imei": "123456789012345"}]
    _logger, _calls = _install(monkeypatch, rows=rows)

    rc = mod.main([])

    assert rc == 10


@pytest.mark.parametrize("imei", ["12345678901234", "1234567890123456", "12345ABCDE12345"])
def test_rejects_invalid_imei(monkeypatch, imei):
    rows = [{"alias": "alias01", "serial": "s1", "imei": imei}]
    _logger, _calls = _install(monkeypatch, rows=rows)

    rc = mod.main([])

    assert rc == 5


def test_accepts_numeric_like_imei_when_excel_reader_already_converted(monkeypatch):
    rows = [{"alias": "alias01", "serial": "s1", "imei": "123456789012345"}]
    _logger, _calls = _install(monkeypatch, rows=rows)

    rc = mod.main([])

    assert rc == 0


def test_detects_duplicate_alias(monkeypatch):
    rows = [
        {"alias": "dup", "serial": "s1", "imei": "123456789012345"},
        {"alias": "dup", "serial": "s2", "imei": "223456789012345"},
    ]
    _logger, _calls = _install(monkeypatch, rows=rows)

    rc = mod.main([])

    assert rc == 11


def test_detects_duplicate_imei(monkeypatch):
    rows = [
        {"alias": "a1", "serial": "s1", "imei": "123456789012345"},
        {"alias": "a2", "serial": "s2", "imei": "123456789012345"},
    ]
    _logger, _calls = _install(monkeypatch, rows=rows)

    rc = mod.main([])

    assert rc == 11


def test_does_not_log_alias_imei_or_serial(monkeypatch):
    alias = "abcdef"
    imei = "123456789012345"
    serial = "serial-secret"
    rows = [{"alias": alias, "serial": serial, "imei": imei}]
    logger, _calls = _install(monkeypatch, rows=rows)

    rc = mod.main([])

    joined = "\n".join(logger.info_messages + logger.error_messages)
    assert rc == 0
    assert alias not in joined
    assert imei not in joined
    assert serial not in joined


def test_returns_8_on_excel_or_sheet_unknown(monkeypatch):
    logger = DummyLogger()
    monkeypatch.setattr(mod, "AppLogger", lambda _base_dir: logger)
    monkeypatch.setattr(mod, "load_config", lambda: {"excel": {"path": "X:/missing.xlsm"}})
    monkeypatch.setattr(mod.Path, "exists", lambda _self: True)

    class BrokenReader:
        def __init__(self, _path):
            return None

        def is_file_open(self):
            return False

        def read_targets(self):
            raise KeyError("sheet not found")

    monkeypatch.setattr(mod, "ExcelReader", BrokenReader)

    rc = mod.main([])

    assert rc == 8


def test_does_not_use_external_system_or_browser_symbols():
    assert not hasattr(mod, "Browser")
    assert not hasattr(mod, "HenngeHandler")
    assert not hasattr(mod, "SmsmHandler")


def test_does_not_write_excel(monkeypatch):
    rows = [{"alias": "alias01", "serial": "s1", "imei": "123456789012345"}]
    _logger, calls = _install(monkeypatch, rows=rows)

    rc = mod.main([])

    assert rc == 0
    assert calls["save"] == 0
    assert not hasattr(mod, "Workbook")


def test_failure_logs_only_stage_and_exception_type_without_message(monkeypatch):
    secret_message = "C:/secret/path alias=alice imei=123456789012345"
    rows = [{"alias": "abxyz", "serial": "s1", "imei": "123456789012345"}]
    logger, _calls = _install(monkeypatch, rows=rows, read_exc=ValueError(secret_message))

    rc = mod.main([])

    joined_err = "\n".join(logger.error_messages)
    joined_info = "\n".join(logger.info_messages)
    assert rc == 1
    assert "failed_stage=read_targets_started" in joined_err
    assert "exception_type=ValueError" in joined_err
    assert "error_category=unknown_value_error" in joined_info
    assert secret_message not in joined_err
    assert secret_message not in joined_info
    assert "alias=alice" not in joined_err
    assert "alias=alice" not in joined_info
    assert "123456789012345" not in joined_err
    assert "123456789012345" not in joined_info


def test_excel_path_is_not_logged(monkeypatch):
    path = "C:/very/secret/folder/live_targets.xlsm"
    rows = [{"alias": "abxyz", "serial": "s1", "imei": "123456789012345"}]
    logger, _calls = _install(monkeypatch, rows=rows, config_path=path)

    rc = mod.main([])

    joined = "\n".join(logger.info_messages + logger.error_messages)
    assert rc == 0
    assert path not in joined
    assert "live_targets.xlsm" not in joined


def test_xlsm_is_allowed(monkeypatch):
    rows = [{"alias": "abxyz", "serial": "s1", "imei": "123456789012345"}]
    _logger, _calls = _install(monkeypatch, rows=rows, config_path="X:/ok.xlsm")

    rc = mod.main([])

    assert rc == 0


@pytest.mark.parametrize(
    "exc, expected_type",
    [
        (PermissionError("permission denied"), "PermissionError"),
        (BadZipFile("bad zip"), "BadZipFile"),
        (mod.InvalidFileException("invalid file"), "InvalidFileException"),
    ],
)
def test_distinguishes_specific_exception_types(monkeypatch, exc, expected_type):
    rows = [{"alias": "abxyz", "serial": "s1", "imei": "123456789012345"}]
    logger, _calls = _install(monkeypatch, rows=rows, read_exc=exc)

    rc = mod.main([])

    joined_err = "\n".join(logger.error_messages)
    assert rc == 1
    assert f"exception_type={expected_type}" in joined_err


def test_other_exception_returns_1(monkeypatch):
    rows = [{"alias": "abxyz", "serial": "s1", "imei": "123456789012345"}]
    logger, _calls = _install(monkeypatch, rows=rows, read_exc=RuntimeError("boom"))

    rc = mod.main([])

    assert rc == 1
    assert "exception_type=RuntimeError" in "\n".join(logger.error_messages)


def test_open_workbook_uses_test_environment_identity(monkeypatch):
    identity_hash = "a" * 64
    monkeypatch.setenv(WEB_IDENTITY_ENV_TEST, identity_hash)
    rows = [{"alias": "alias01", "serial": "s1", "imei": "123456789012345"}]
    _logger, calls = _install(monkeypatch, rows=rows, is_open=True)

    rc = mod.main([])

    assert rc == 0
    assert calls["detection_kwargs"]["configured_web_identity_hash"] == identity_hash
    assert calls["detection_kwargs"]["web_identity_source"] == "environment"
    assert calls["detection_kwargs"]["web_identity_valid"] is True
    assert calls["detection_kwargs"]["web_identity_mode"] == "test"
    assert calls["save_kwargs"]["configured_web_identity_hash"] == identity_hash
    assert calls["save_kwargs"]["web_identity_mode"] == "test"


def test_unmatched_workbook_skips_save_close_read_and_reopen(monkeypatch):
    rows = [{"alias": "alias01", "serial": "s1", "imei": "123456789012345"}]
    _logger, calls = _install(monkeypatch, rows=rows, is_open=True, detection_match=False)

    rc = mod.main([])

    assert rc == 9
    assert calls["save"] == 0
    assert calls["close"] == 0
    assert calls["read_targets"] == 0
    assert calls["reopen"] == 0


def test_save_failure_skips_close_unlock_and_read(monkeypatch):
    rows = [{"alias": "alias01", "serial": "s1", "imei": "123456789012345"}]
    _logger, calls = _install(
        monkeypatch,
        rows=rows,
        is_open=True,
        save_exc=mod.SaveCloseWorkbookError("secret"),
    )

    rc = mod.main([])

    assert rc == 12
    assert calls["save"] == 1
    assert calls["close"] == 0
    assert calls["unlock"] == 0
    assert calls["read_targets"] == 0


def test_close_failure_skips_unlock_and_read(monkeypatch):
    rows = [{"alias": "alias01", "serial": "s1", "imei": "123456789012345"}]
    _logger, calls = _install(
        monkeypatch,
        rows=rows,
        is_open=True,
        close_exc=mod.SaveCloseWorkbookError("secret"),
    )

    rc = mod.main([])

    assert rc == 12
    assert calls["save"] == 1
    assert calls["close"] == 1
    assert calls["unlock"] == 0
    assert calls["read_targets"] == 0


def test_unlock_failure_skips_read_and_reopens_once(monkeypatch):
    rows = [{"alias": "alias01", "serial": "s1", "imei": "123456789012345"}]
    _logger, calls = _install(
        monkeypatch,
        rows=rows,
        is_open=True,
        unlock_exc=mod.UnlockTimeoutError("secret"),
    )

    rc = mod.main([])

    assert rc == 13
    assert calls["read_targets"] == 0
    assert calls["reopen"] == 1


def test_reopen_failure_returns_14_and_is_attempted_once(monkeypatch):
    rows = [{"alias": "alias01", "serial": "s1", "imei": "123456789012345"}]
    _logger, calls = _install(monkeypatch, rows=rows, is_open=True, reopen_fails=True)

    rc = mod.main([])

    assert rc == 14
    assert calls["read_targets"] == 1
    assert calls["reopen"] == 1


def test_missing_imei_returns_10(monkeypatch):
    rows = [{"alias": "alias01", "serial": "s1", "imei": ""}]
    _logger, _calls = _install(monkeypatch, rows=rows)

    assert mod.main([]) == 10


def test_non_dictionary_row_returns_11(monkeypatch):
    _logger, _calls = _install(monkeypatch, rows=[("alias01", "123456789012345")])

    assert mod.main([]) == 11


# ---------------------------------------------------------------------------
# ValueError 分類．ログ安全性
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("message, expected_category, expected_code", [
    ("Excel 4行目の必須項目が不足: serial",  "missing_required_field", 10),
    ("IMEIが整数ではありません: 1.5",         "unknown_value_error",      1),
    ("IMEIに数字以外が含まれています: ABC",  "unknown_value_error",      1),
    ("IMEIは15桁ではありません",                "invalid_imei",            5),
    ("その他のエラー",                                 "unknown_value_error",      1),
])
def test_value_error_category_and_code(monkeypatch, message, expected_category, expected_code):
    rows = [{"alias": "alias01", "serial": "s1", "imei": "123456789012345"}]
    logger, _calls = _install(monkeypatch, rows=rows, read_exc=ValueError(message))

    rc = mod.main([])

    joined_info = "\n".join(logger.info_messages)
    joined_err = "\n".join(logger.error_messages)
    assert rc == expected_code
    assert f"error_category={expected_category}" in joined_info
    assert "exception_type=ValueError" in joined_err
    assert message not in joined_info
    assert message not in joined_err


def test_value_error_does_not_log_row_number(monkeypatch):
    row_hint = "4行目"
    rows = [{"alias": "alias01", "serial": "s1", "imei": "123456789012345"}]
    logger, _calls = _install(monkeypatch, rows=rows, read_exc=ValueError(f"Excel {row_hint}の必須項目が不足: serial"))

    mod.main([])

    joined = "\n".join(logger.info_messages + logger.error_messages)
    assert row_hint not in joined


def test_value_error_during_read_still_reopens_excel_once(monkeypatch):
    rows = [{"alias": "alias01", "serial": "s1", "imei": "123456789012345"}]
    _logger, calls = _install(
        monkeypatch,
        rows=rows,
        is_open=True,
        read_exc=ValueError("IMEIは15桁ではありません"),
    )

    rc = mod.main([])

    assert rc == 5
    assert calls["read_targets"] == 1
    assert calls["reopen"] == 1


def test_multiple_valid_rows_return_0(monkeypatch):
    rows = [
        {"alias": "a1", "serial": "s1", "imei": "111111111111111"},
        {"alias": "a2", "serial": "s2", "imei": "222222222222222"},
    ]
    _logger, calls = _install(monkeypatch, rows=rows)

    rc = mod.main([])

    assert rc == 0
    assert calls["read_targets"] == 1

def test_multiple_valid_rows_alias_imei_pair_maintained(monkeypatch):
    observed = {}

    original_install = _install

    rows = [
        {"alias": "pairA", "serial": "sA", "imei": "111111111111111"},
        {"alias": "pairB", "serial": "sB", "imei": "222222222222222"},
    ]

    class ObservingReader(FakeReader):
        def read_targets(self):
            result = super().read_targets()
            observed["rows"] = result
            return result

    logger = DummyLogger()
    calls = {"load_config": 0, "reader_ctor": 0, "reader_path": "", "is_file_open": 0, "read_targets": 0,
             "detect": 0, "save": 0, "close": 0, "unlock": 0, "reopen": 0, "events": [],
             "detection_kwargs": {}, "save_kwargs": {}}

    monkeypatch.setattr(mod, "AppLogger", lambda _base_dir: logger)
    monkeypatch.setattr(mod, "load_config", lambda: {"excel": {"path": "X:/t.xlsm"}})
    monkeypatch.setattr(mod, "ExcelReader",
                        lambda fp: ObservingReader(fp, rows, calls, is_open=False))
    monkeypatch.setattr(mod.Path, "exists", lambda _self: True)
    monkeypatch.setattr(mod, "detect_target_workbook", lambda *a, **kw: None)
    monkeypatch.setattr(mod, "save_and_close_target_workbook", lambda *a, **kw: None)
    monkeypatch.setattr(mod, "_wait_unlock", lambda *a, **kw: None)
    monkeypatch.setattr(mod, "reopen_excel", lambda *a, **kw: None)

    mod.main([])

    assert observed["rows"][0]["alias"] == "pairA"
    assert observed["rows"][0]["imei"] == "111111111111111"
    assert observed["rows"][1]["alias"] == "pairB"
    assert observed["rows"][1]["imei"] == "222222222222222"


@pytest.mark.parametrize("separator", [" ", "\u3000", "\u00a0", "\u2009", "\u202f", "\u200b"])
def test_imei_format_normalizes_allowed_separators(separator):
    normalized, category = mod._normalize_imei_for_diagnosis(
        f"35{separator}936730{separator}687217{separator}7"
    )

    assert normalized == "359367306872177"
    assert category == "valid_after_whitespace_normalization_count"


@pytest.mark.parametrize(
    "value, expected_category",
    [
        ("12345678901234", "invalid_length_count"),
        ("1234567890123456", "invalid_length_count"),
        ("１２３４５６７８９０１２３４５", "non_ascii_digit_count"),
        ("123456789012-45", "non_digit_character_count"),
        (12345678901234.5, "non_integer_numeric_count"),
    ],
)
def test_imei_format_classifies_invalid_values(value, expected_category):
    _normalized, category = mod._normalize_imei_for_diagnosis(value)

    assert category == expected_category


def test_collect_imei_format_counts_scans_all_rows_without_writing(tmp_path: Path):
    workbook_path = tmp_path / "targets.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "HENNGE登録作業必要情報"
    sheet.append(["", "", "alias", "serial", "imei"])
    sheet.append(["", "", "", "", ""])
    sheet.append(["", "", "", "", ""])
    sheet.append(["", "", "a1", "s1", "123456789012345"])
    sheet.append(["", "", "a2", "s2", "12345678901234"])
    sheet.append(["", "", "a3", "s3", "1234567890123456"])
    sheet.append(["", "", "a4", "s4", "１２３４５６７８９０１２３４５"])
    sheet.append(["", "", "a5", "s5", "123456789012-45"])
    sheet.append(["", "", "a6", "s6", 12345678901234.5])
    sheet.append(["", "", "a7", "s7", "35 936730 687217 7"])
    workbook.save(workbook_path)
    workbook.close()

    counts = mod._collect_imei_format_counts(workbook_path)

    assert counts["total_candidate_row_count"] == 7
    assert counts["valid_original_count"] == 1
    assert counts["valid_after_whitespace_normalization_count"] == 1
    assert counts["invalid_length_count"] == 2
    assert counts["non_ascii_digit_count"] == 1
    assert counts["non_digit_character_count"] == 1
    assert counts["non_integer_numeric_count"] == 1
    assert workbook_path.exists()


def test_normal_read_and_format_scan_use_same_file_and_candidate_row(tmp_path: Path):
    workbook_path = tmp_path / "same-target.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "HENNGE登録作業必要情報"
    sheet.cell(4, 3, "alias01")
    sheet.cell(4, 4, "serial01")
    sheet.cell(4, 5, "35 936730 687217 7")
    workbook.save(workbook_path)
    workbook.close()

    rows = ExcelReader(str(workbook_path)).read_targets()
    counts = mod._collect_imei_format_counts(workbook_path)

    assert counts["total_candidate_row_count"] == 1
    assert counts["valid_after_whitespace_normalization_count"] == 1
    assert rows == [{"alias": "alias01", "serial": "serial01", "imei": "359367306872177"}]


def test_imei_format_mode_does_not_call_read_targets_and_logs_only_counts(monkeypatch):
    logger, calls = _install(
        monkeypatch,
        rows=[{"alias": "secret-alias", "serial": "secret-serial", "imei": "secret-imei"}],
    )
    counts = mod._new_imei_format_counts()
    counts["total_candidate_row_count"] = 2
    counts["valid_original_count"] = 1
    counts["invalid_length_count"] = 1
    monkeypatch.setattr(mod, "_collect_imei_format_counts", lambda _path: counts)

    rc = mod.main(["--diagnose-imei-format"])

    assert rc == 5
    assert calls["read_targets"] == 0
    keys = {message.split("=", 1)[0] for message in logger.info_messages}
    assert keys == set(mod.IMEI_FORMAT_COUNTERS)
    joined = "\n".join(logger.info_messages + logger.error_messages)
    assert "secret-alias" not in joined
    assert "secret-serial" not in joined
    assert "secret-imei" not in joined
